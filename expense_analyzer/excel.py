from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Sequence

from .payee_map import format_label_from_cells, normalize_label, payees_for_label
from .payments_map import PaymentRule
from .models import Transaction
from .summary import CollectionSummaryRow


def _norm(text: str) -> str:
    # Normalise payee text for lookups: lowercase alnum and single spaces
    import re

    t = (text or "").lower()
    t = re.sub(r"[^a-z0-9\s]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


@dataclass(frozen=True)
class PaidSection:
    block_col: int
    flat_col: int
    paid_cols: Sequence[int]


def _find_collection_sheet(wb) -> tuple[object, int]:
    """Locate the collection summary sheet and header row."""

    target_sheet = None
    for ws in wb.worksheets:
        key = _norm(ws.title).replace(" ", "")
        if key in {"collectionsummary", "collectionsummarry"}:
            target_sheet = ws
            break
    if target_sheet is None:
        target_sheet = wb.worksheets[0]

    header_row_idx = None
    for r in range(1, min(10, target_sheet.max_row) + 1):
        row_vals = [target_sheet.cell(row=r, column=c).value for c in range(1, target_sheet.max_column + 1)]
        if any(isinstance(v, str) and "paid" in v.lower() for v in row_vals):
            header_row_idx = r
            break
    if header_row_idx is None:
        raise ValueError("Could not locate a header row with a 'Paid' column in the workbook")

    return target_sheet, header_row_idx


def _find_sheet_by_name(wb, name: str):
    target = _norm(name)
    for ws in wb.worksheets:
        if _norm(ws.title) == target:
            return ws
    return None



def _apply_account_adjustments(account_sheet, template_account_sheet) -> None:
    if account_sheet is None:
        return

    # Clear E5:E21 in the Account sheet
    for row in range(5, 22):
        account_sheet.cell(row=row, column=5).value = None

    if template_account_sheet is None:
        return

    # Copy Account!C31 -> Account!C4 using template values
    account_sheet["C4"].value = template_account_sheet["C31"].value

    # Copy Account!C56:I56 -> Account!C36:I36 using template values
    for column in range(3, 10):  # Columns C (3) through I (9)
        account_sheet.cell(row=36, column=column).value = template_account_sheet.cell(row=56, column=column).value


def _calculate_payment_totals(transactions: Sequence[Transaction], rules: Sequence[PaymentRule] | None) -> Dict[str, float]:
    totals: Dict[str, float] = {}
    if not rules:
        return totals
    for rule in rules:
        total = 0.0
        for tx in transactions:
            if rule.matches(tx):
                total += abs(tx.amount)
        if total > 0:
            totals[_norm(rule.row_label)] = round(total, 2)
    return totals



def _apply_payment_totals(account_sheet, totals: Mapping[str, float]) -> None:
    if account_sheet is None or not totals:
        return
    for row_idx in range(1, account_sheet.max_row + 1):
        label = account_sheet.cell(row=row_idx, column=2).value
        if not isinstance(label, str):
            continue
        key = _norm(label)
        if key in totals:
            account_sheet.cell(row=row_idx, column=5).value = totals[key]


def _detect_sections(ws, header_row_idx: int) -> List[PaidSection]:
    sections: List[PaidSection] = []
    col = 1
    while col <= ws.max_column:
        header = ws.cell(row=header_row_idx, column=col).value
        if isinstance(header, str) and header.strip().lower() == "block":
            block_col = col
            flat_col = col + 1 if col + 1 <= ws.max_column else None
            scan_col = col
            paid_cols: List[int] = []
            while scan_col <= ws.max_column:
                value = ws.cell(row=header_row_idx, column=scan_col).value
                if scan_col != col and isinstance(value, str) and value.strip().lower() == "block":
                    break
                if isinstance(value, str) and "paid" in value.lower():
                    paid_cols.append(scan_col)
                scan_col += 1
            if flat_col is not None and paid_cols:
                sections.append(PaidSection(block_col, flat_col, tuple(paid_cols)))
            col = scan_col
        else:
            col += 1
    if not sections:
        raise ValueError("Could not locate block/flat columns with Paid headers")
    return sections


def _build_receipts_by_payee(collection_rows: Iterable[CollectionSummaryRow]) -> Dict[str, float]:
    totals: Dict[str, float] = {}
    for row in collection_rows:
        name = (row.payee_payer or "").strip()
        if not name:
            continue
        key = _norm(name)
        totals[key] = totals.get(key, 0.0) + float(row.receipts or 0.0)
    return totals


def _sum_for_payees(payees: Sequence[str], receipts_by_payee: Mapping[str, float]) -> float:
    total = 0.0
    for payee in payees:
        total += receipts_by_payee.get(_norm(payee), 0.0)
    return total


def update_paid_columns(
    template_path: Path,
    output_path: Path,
    collection_rows: Iterable[CollectionSummaryRow],
    period_transactions: Sequence[Transaction] | None = None,
    payee_map: Mapping[str, Sequence[str]] | None = None,
    payment_rules: Sequence[PaymentRule] | None = None,
    fixed_paid_overrides: Mapping[str, float] | None = None,
    period_start: date | None = None,
    period_end: date | None = None,
) -> None:
    """
    Update only the 'Paid' columns of the 'Collection Summary' sheet using
    a supplied Payee <-> Block/Flat mapping and save to ``output_path``.
    """

    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency guidance
        raise SystemExit(
            "openpyxl is required for Excel output. Install with: pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=str(template_path))
    template_values_wb = load_workbook(filename=str(template_path), data_only=True)
    ws, header_row_idx = _find_collection_sheet(wb)
    sections = _detect_sections(ws, header_row_idx)
    payee_map = payee_map or {}
    fixed_paid_overrides = {
        normalize_label(label): float(amount)
        for label, amount in (fixed_paid_overrides or {}).items()
        if amount is not None
    }

    receipts_by_payee = _build_receipts_by_payee(collection_rows)
    payment_totals = _calculate_payment_totals(period_transactions or [], payment_rules)

    for section in sections:
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            label = format_label_from_cells(
                ws.cell(row=row_idx, column=section.block_col).value,
                ws.cell(row=row_idx, column=section.flat_col).value,
            )
            if not label:
                continue
            label_key = normalize_label(label)
            if fixed_paid_overrides and label_key in fixed_paid_overrides:
                value = round(float(fixed_paid_overrides[label_key]), 2)
            else:
                payees = payees_for_label(payee_map, label)
                value = None
                if payees:
                    amount = _sum_for_payees(payees, receipts_by_payee)
                    if amount > 0:
                        value = round(float(amount), 2)
            for col_idx in section.paid_cols:
                ws.cell(row=row_idx, column=col_idx).value = value

    if period_end:
        ws.cell(row=1, column=3).value = period_end

    account_sheet = _find_sheet_by_name(wb, "account")
    template_account_sheet = _find_sheet_by_name(template_values_wb, "account")
    _apply_account_adjustments(account_sheet, template_account_sheet)
    _apply_payment_totals(account_sheet, payment_totals)
    if account_sheet is not None:
        if period_start:
            account_sheet["A4"].value = period_start
        if period_end:
            account_sheet["A31"].value = period_end

    wb.save(str(output_path))
